#!/usr/bin/python
# !coding=utf-8


import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import LineChart, AreaChart
from openpyxl.chart import Reference, Series


class OprateExcel(object):
    """
    这个类的作用是用来实现对excel表的操作
    对于这个类我们提出下面几个要求
    1.获得excel操作对象
    2.创建excel表--->根据指定对象
    3.读取指定行的数据
    4,遍历

    对于这个类的使用方法是
    1.进行初始化操作，获得操作对象
    2.如果是读取一个文件，调用openFile
    3.通过filerow来获得文件中的行数，通过
    4.通过readFileRow()传递对应的索引获得数据

    """
    wb = None
    excelName = ""
    ws = None

    def __init__(self, excelName):
        self.excelName = excelName
        print("文件路劲", excelName)

    def saveFile(self):
        self.wb.save(self.excelName)

    def writeFile(self, rows):
        index = self.ws.max_row + 1
        for column in range(len(rows)):
            cell = self.ws.cell(row=index, column=column + 1)
            cell.value = rows[column]

        pass

    def writeInit(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def openFile(self):
        """
        参数是一个回调函数，代表的是文件是否存在的
        :param callback:
        :return:
        """

        self.wb = load_workbook(self.excelName)
        self.ws = self.wb.active

    def fileIsExist(self):
        print(self.excelName)
        value = os.path.exists(self.excelName)
        return value

    def fileRows(self):
        print("行", self.ws)  # 这行出错
        return self.ws.max_row

    def readFileRow(self, index):
        """
        读取文件指定行的数据，传递的是文件行数的索引值，index从0开始
        :param index:
        :return:
        """
        list1 = list()
        for cell in list(self.ws.rows)[index]:
            list1.append(cell.value)
        return list1
